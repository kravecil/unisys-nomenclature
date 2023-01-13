from django.core.management import BaseCommand
from django.db import transaction, IntegrityError
from django.conf import settings
import datetime
from openpyxl import load_workbook
from pathlib import Path

from nomenclature.models.requests import Request
from nomenclature.models.products import Product
from nomenclature.models.approvers import Approver
from nomenclature.models.units import Unit
from nomenclature.models.okpd import Okpd
from nomenclature.models.okved import Okved
from nomenclature.models.ens import Ens

class Command(BaseCommand):
    SOURCE_DIR = settings.BASE_DIR / 'import_source'
    
    def add_arguments(self, parser):
        parser.add_argument('--approvers')
        parser.add_argument('--mdb')
        parser.add_argument('--units')
        parser.add_argument('--okpd')
        parser.add_argument('--okved')
        parser.add_argument('--ens')
        parser.add_argument('--all')

    @transaction.atomic
    def handle(self, *args, **options):
        print('НАЧИНАЕМ ИМПОРТИРОВАНИЕ:')
        print('-----------------------')

        tr = transaction.savepoint()
        try:
            if options['all'] or options['approvers']:
                # Создание польователей, согласующих заявки
                print('Создаём визирующих...')
                # Approver.objects.create(auth_id=79) # Александрова Ольга Алексеевна
                # Approver.objects.create(auth_id=167) # Волгина Карине Владимировна
                # Approver.objects.create(auth_id=83) # Мухина Инна Владимировна
                Approver.objects.create(auth_id=2) # Кравец Илья Владимирович
                Approver.objects.create(auth_id=3) # Галлямов Марат Асгатович
                Approver.objects.create(auth_id=452) # Вафина Зайтуна Ахтямовна

            if options['all'] or options['mdb']:
                # Импортируем из базы Инира
                print('Импортируем ТМЦ из базы MDB...')
                print('...загружаем книгу Excel')
                wb = load_workbook(self.SOURCE_DIR / 'tmc.xlsx')
                sheet = wb.active

                print('...добавляем записи')
                row = 1
                while True:
                    if not sheet.cell(column=1, row=row).value: break
                    
                    # номенклатурный номер
                    number = str(sheet.cell(column=1, row=row).value or '').strip()
                    # полное наименование
                    name = str(sheet.cell(column=4, row=row).value or '').strip()
                    # единица измерения
                    unit = str(sheet.cell(column=27, row=row).value or '').strip().replace('.', '')
                    if unit:
                        units = Unit.objects.filter(shortname=unit)
                        if len(units)==0:
                            units = Unit.objects.filter(code=unit)
                    unit = units.first()
                    # технические характеристики
                    notes = ''

                    data = str(sheet.cell(column=11, row=row).value or '').strip().replace('.', '')
                    if data: notes += 'Размер 1: %s\n' % data
                    data = str(sheet.cell(column=12, row=row).value or '').strip().replace('.', '')
                    if data: notes += 'Размер 2: %s\n' % data
                    data = str(sheet.cell(column=13, row=row).value or '').strip().replace('.', '')
                    if data: notes += 'Размер 3: %s\n' % data
                    data = str(sheet.cell(column=15, row=row).value or '').strip().replace('.', '')
                    if data: notes += 'Характеристика: %s\n' % data
                    data = str(sheet.cell(column=20, row=row).value or '').strip().replace('.', '')
                    if data: notes += 'ГОСТ/Сортамент: %s\n' % data
                    data = str(sheet.cell(column=21, row=row).value or '').strip().replace('.', '')
                    if data: notes += 'ГОСТ/Марка: %s\n' % data

                    okpd = str(sheet.cell(column=16, row=row).value or '').strip()
                    okpd_obj = None
                    okved = str(sheet.cell(column=19, row=row).value or '').strip()
                    okved_obj = None

                    okpd_queryset = Okpd.objects.filter(code=okpd)
                    if okpd_queryset.count() > 0: okpd_obj = okpd_queryset.first()
                    okved_queryset = Okved.objects.filter(code=okved)
                    if okved_queryset.count() > 0: okved_obj = okved_queryset.first()

                    notes += 'Импортировано из mdb!'
                    #дата добавления
                    created_at = None
                    date = str(sheet.cell(column=10, row=row).value or '').strip().replace('.', '')
                    if date:
                        date = date.split(' ')
                        try:
                            created_at = datetime.datetime.strptime(date[0], '%m/%d/%y')
                        except:
                            pass
                    
                    product = Product.objects.create(
                        number=number,
                        name=name,
                        notes=notes,
                        okpd=okpd_obj,
                        okved=okved_obj,
                        unit=unit)
                    Request.objects.create(
                        created_at=created_at or datetime.datetime.now(),
                        product=product,
                        approved_at=datetime.datetime.now(),
                        person_who_created=249, # Абрахманов Инир Кашфуллович
                        )
                    
                    row += 1

                wb.close()

            if options['all'] or options['units']:
                # Заполняем единицы измерения
                print('Заполняем справочник единиц измерения...')
                print('...загружаем книгу Excel')
                wb = load_workbook(self.SOURCE_DIR / 'okei.xlsx')
                sheet = wb.active

                print('...добавляем записи')
                row = 1
                while True:
                    if not sheet.cell(column=1, row=row).value: break

                    Unit.objects.create(
                        code = str(sheet.cell(column=1, row=row).value or '').strip(),
                        name = str(sheet.cell(column=2, row=row).value or '').strip(),
                        shortname = str(sheet.cell(column=3, row=row).value or '').strip(),
                        codename = str(sheet.cell(column=4, row=row).value or '').strip(),
                        shortname_international = str(sheet.cell(column=5, row=row).value or '').strip(),
                        codename_international = str(sheet.cell(column=6, row=row).value or '').strip(),
                    )
                    
                    row += 1

                wb.close()

            if options['all'] or options['okpd']:
                # Заполняем ОКПД2
                print('Заполняем данные ОКПД2...')
                print('...загружаем книгу Excel')
                wb = load_workbook(self.SOURCE_DIR / 'okpd2.xlsx')
                sheet = wb.active

                print('...добавляем записи')
                row = 1
                while True:
                    if not sheet.cell(column=1, row=row).value: break

                    code = str(sheet.cell(column=1, row=row).value or '').strip()
                    code = code.replace(',', '.')

                    Okpd.objects.get_or_create(code=code, defaults={
                        'name': str(sheet.cell(column=2, row=row).value or '').strip(),
                    })
                    
                    row += 1

                wb.close()

            if options['all'] or options['okved']:
                # Заполняем ОКВЭД2
                print('Заполняем данные ОКВЭД2...')
                print('...загружаем книгу Excel')
                wb = load_workbook(self.SOURCE_DIR / 'okved2.xlsx')
                sheet = wb.active

                print('...добавляем записи')
                row = 1
                while True:
                    if not sheet.cell(column=1, row=row).value: break

                    Okved.objects.create(
                        code = str(sheet.cell(column=1, row=row).value or '').strip(),
                        name = str(sheet.cell(column=2, row=row).value or '').strip(),
                    )
                    
                    row += 1

                wb.close()

            if options['all'] or options['ens']:
                # Загружаем данные из ЕНС ОКПД2
                print('Заполняем данные из ЕНС...')
                print('...загружаем книгу Excel')
                wb = load_workbook(self.SOURCE_DIR / 'ens.xlsx')
                sheet = wb.active

                print('...добавляем записи')
                row = 1
                while True:
                    if not sheet.cell(column=1, row=row).value: break

                    unit = str(sheet.cell(column=3, row=row).value or '').strip()
                    if unit:
                        number = str(sheet.cell(column=2, row=row).value or '').strip()
                        notes = str(sheet.cell(column=9, row=row).value or '').strip()
                        notes = notes.replace(';', ';\n')
                        # TODO: исправить пробел пи импортировании, если он есть после ;

                        Ens.objects.get_or_create(number=number, defaults={
                            'name': str(sheet.cell(column=1, row=row).value or '').strip(),
                            'unit': unit,
                            'okpd': str(sheet.cell(column=7, row=row).value or '').strip(),
                            'okved': str(sheet.cell(column=6, row=row).value or '').strip(),
                            'notes': notes,
                        })
                    
                    row += 1

                wb.close()

            transaction.savepoint_commit(tr)
        except Exception as e:
            print(e)
            transaction.savepoint_rollback(tr)

        print('-------------------------')
        print('ИМПОРТИРОВАНИЕ ЗАКОНЧЕНО!')